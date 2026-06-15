from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from utils.tlg_data_cleaning import MONTHS, load_tlg_trial_balance, normalize_text
from utils.tlg_financial_statements import prepare_tlg_detail


TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "Informes_mensualizados_template.xlsx"
PYG_YEAR_STARTS = {2024: 5, 2025: 18, 2026: 31}
PYG_YEAR_TOTALS = {2024: 17, 2025: 30, 2026: 43}


def _period_from_metadata(metadata: dict[str, str | None]) -> tuple[int, int]:
    month_name = str(metadata.get("mes") or "").strip().lower()
    year_text = str(metadata.get("anio") or "").strip()
    month = MONTHS.get(month_name)
    if month is None or not year_text.isdigit():
        raise ValueError(
            "No fue posible identificar el mes y el año del balance. "
            "Verifica que el encabezado indique el periodo."
        )
    return int(year_text), month


def _account4_values(detail: pd.DataFrame) -> tuple[dict[str, float], dict[str, float]]:
    data = detail.copy()
    data["CUENTA_4"] = (
        data["CODIGO_CUENTA"]
        .fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)
        .str[:4]
    )
    data = data[data["CUENTA_4"].str.len() == 4].copy()

    balance = data.groupby("CUENTA_4")["SALDO_FINAL"].sum().to_dict()
    # The template carries the accumulated result in account 3605.
    balance["3605"] = float(
        data.loc[data["CLASE"].isin(["4", "5", "6", "7"]), "SALDO_FINAL"].sum()
    )
    data["VALOR_PYG"] = data["MOVIMIENTO_DEBITO"] - data["MOVIMIENTO_CREDITO"]
    income = data["CLASE"] == "4"
    data.loc[income, "VALOR_PYG"] = (
        data.loc[income, "MOVIMIENTO_CREDITO"]
        - data.loc[income, "MOVIMIENTO_DEBITO"]
    )
    pyg = data.groupby("CUENTA_4")["VALOR_PYG"].sum().to_dict()
    return balance, pyg


def _find_bce_month_column(worksheet, year: int, month: int) -> int:
    for column in range(1, min(worksheet.max_column, 48) + 1):
        value = worksheet.cell(2, column).value
        if isinstance(value, (datetime, pd.Timestamp)):
            if value.year == year and value.month == month:
                return column
        if month == 12 and isinstance(value, str):
            normalized = value.lower().replace(" ", "")
            if normalized.startswith("saldofinal") and str(year) in normalized:
                return column
    raise ValueError(
        f"La plantilla no contiene el periodo {month:02d}/{year} en la hoja BCE."
    )


def _find_pyg_month_column(year: int, month: int) -> int:
    if year not in PYG_YEAR_STARTS:
        raise ValueError(f"La plantilla no contiene el año {year} en la hoja P Y G.")
    return PYG_YEAR_STARTS[year] + month - 1


def _write_mapped_accounts(worksheet, column: int, values: dict[str, float]) -> int:
    updated = 0
    for row in range(1, worksheet.max_row + 1):
        raw_code = worksheet.cell(row, 3).value
        if raw_code is None:
            continue
        code = str(raw_code).replace(".0", "").strip()
        if not code.isdigit() or len(code) != 4:
            continue
        target_cell = worksheet.cell(row, column)
        if isinstance(target_cell, MergedCell):
            continue
        target_cell.value = float(values.get(code, 0.0))
        updated += 1
    return updated


def _replace_external_formulas(workbook, cached_workbook) -> None:
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        cached_sheet = cached_workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if (
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                    and "[" in cell.value
                ):
                    cached_value = cached_sheet[cell.coordinate].value
                    cell.value = cached_value if cached_value is not None else 0
    workbook._external_links = []


def _configure_visible_periods(
    workbook,
    start_year: int,
    monthly_periods: list[tuple[int, int]],
    preserve_existing: bool = False,
) -> None:
    bce = workbook["BCE"]
    pyg = workbook["P Y G"]

    base_bce_column = _find_bce_month_column(bce, start_year, 12)
    visible_bce = {4, base_bce_column}
    visible_pyg = {4, PYG_YEAR_TOTALS[start_year]}

    if preserve_existing:
        visible_bce.update(
            column
            for column in range(5, 49)
            if not bce.column_dimensions[get_column_letter(column)].hidden
        )
        visible_pyg.update(
            column
            for column in range(5, 44)
            if not pyg.column_dimensions[get_column_letter(column)].hidden
        )

    for year, month in monthly_periods:
        visible_bce.add(_find_bce_month_column(bce, year, month))
        visible_pyg.add(_find_pyg_month_column(year, month))

    for column in range(5, 49):
        bce.column_dimensions[get_column_letter(column)].hidden = column not in visible_bce
    for column in range(5, 44):
        pyg.column_dimensions[get_column_letter(column)].hidden = column not in visible_pyg

    bce_header = bce.cell(2, base_bce_column)
    if not isinstance(bce_header, MergedCell):
        bce_header.value = f"Saldo final {start_year}"
    pyg_header = pyg.cell(3, PYG_YEAR_TOTALS[start_year])
    if not isinstance(pyg_header, MergedCell):
        pyg_header.value = f"Saldo final {start_year}"

    bce.sheet_state = "visible"
    pyg.sheet_state = "visible"
    workbook.active = workbook.index(bce)


def _load_template(previous_file: BinaryIO | None):
    if previous_file is not None:
        previous_file.seek(0)
        source_bytes = previous_file.read()
        source = BytesIO(source_bytes)
        cached_source = BytesIO(source_bytes)
        source_name = "Informe mensualizado anterior"
    else:
        if not TEMPLATE_PATH.exists():
            raise FileNotFoundError("No se encontró la plantilla oficial mensualizada.")
        source = TEMPLATE_PATH
        cached_source = TEMPLATE_PATH
        source_name = "Plantilla oficial"

    workbook = load_workbook(source, data_only=False, keep_links=False)
    cached_workbook = load_workbook(cached_source, data_only=True, keep_links=False)
    missing = {"BCE", "P Y G"} - set(workbook.sheetnames)
    if missing:
        raise ValueError("La plantilla debe contener las hojas BCE y P Y G.")
    _replace_external_formulas(workbook, cached_workbook)
    return workbook, source_name


def _infer_start_year(workbook) -> int:
    bce = workbook["BCE"]
    for column in range(5, 49):
        value = bce.cell(2, column).value
        if isinstance(value, str) and value.lower().startswith("saldo final"):
            year_text = "".join(character for character in value if character.isdigit())
            if len(year_text) == 4:
                return int(year_text)
    for year in sorted(PYG_YEAR_TOTALS, reverse=True):
        try:
            column = _find_bce_month_column(bce, year, 12)
        except ValueError:
            continue
        if not bce.column_dimensions[get_column_letter(column)].hidden:
            return year
    return 2025


def _period_metrics(
    detail: pd.DataFrame,
    year: int,
    month: int,
    label: str,
    kind: str,
) -> dict[str, object]:
    codes = detail["CODIGO_CUENTA"].astype(str)

    def movement(prefixes: tuple[str, ...], income_sign: bool = False) -> float:
        selected = detail[codes.str.startswith(prefixes)]
        if income_sign:
            return float(
                selected["MOVIMIENTO_CREDITO"].sum()
                - selected["MOVIMIENTO_DEBITO"].sum()
            )
        return float(
            selected["MOVIMIENTO_DEBITO"].sum()
            - selected["MOVIMIENTO_CREDITO"].sum()
        )

    operating_income = movement(("41",), income_sign=True)
    other_income = movement(("42", "47"), income_sign=True)
    direct_costs = movement(("61", "62", "71", "72", "73"))
    administration = movement(("51",))
    selling = movement(("52",))
    other_expenses = movement(("53", "54"))
    income = operating_income + other_income
    costs = direct_costs + administration + selling + other_expenses
    gross_profit = operating_income - direct_costs
    operating_profit = gross_profit - administration - selling
    net_result = income - costs
    depreciation = movement(("5160", "5165", "5260"))
    assets = detail.loc[detail["CLASE"] == "1", "SALDO_FINAL"].sum()
    liabilities = abs(detail.loc[detail["CLASE"] == "2", "SALDO_FINAL"].sum())
    equity = abs(detail.loc[detail["CLASE"] == "3", "SALDO_FINAL"].sum())
    accumulated_result = detail.loc[
        detail["CLASE"].isin(["4", "5", "6", "7"]), "SALDO_FINAL"
    ].sum()
    balance_difference = assets - (liabilities + equity - accumulated_result)
    current_assets = detail.loc[
        codes.str.startswith(("11", "13", "14")), "SALDO_FINAL"
    ].sum()
    current_liabilities = abs(
        detail.loc[codes.str.startswith(("21", "22", "23", "24", "25", "26", "27", "28")), "SALDO_FINAL"].sum()
    )
    return {
        "Tipo": kind,
        "Periodo": label,
        "Año": year,
        "Mes": month,
        "Ingresos": float(income),
        "Ventas": float(operating_income),
        "Utilidad bruta": float(gross_profit),
        "Utilidad operacional": float(operating_profit),
        "Costos y gastos": float(costs),
        "Resultado": float(net_result),
        "EBITDA": float(operating_profit + depreciation),
        "Activos": float(assets),
        "Pasivos": float(liabilities),
        "Patrimonio": float(equity),
        "Activos corrientes": float(current_assets),
        "Pasivos corrientes": float(current_liabilities),
        "Resultado acumulado": float(accumulated_result),
        "Diferencia de cuadre": float(balance_difference),
    }


def _transactional_rows(raw_df: pd.DataFrame) -> pd.DataFrame:
    transactional = raw_df["TRANSACCIONAL"].map(normalize_text).eq("SI")
    detail = raw_df[transactional].copy()
    if detail.empty:
        max_length = raw_df["CODIGO_CUENTA"].astype(str).str.len().max()
        detail = raw_df[
            raw_df["CODIGO_CUENTA"].astype(str).str.len().eq(max_length)
        ].copy()
    return detail


def _third_party_tables(raw_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    detail = _transactional_rows(raw_df)
    detail["CUENTA"] = detail["CODIGO_CUENTA"].astype(str)
    detail["Tercero"] = (
        detail["NOMBRE_TERCERO"].fillna("").astype(str).str.strip()
    )
    detail.loc[detail["Tercero"].eq(""), "Tercero"] = "Sin tercero"
    detail["Identificación"] = (
        detail["IDENTIFICACION"].fillna("").astype(str).str.replace(r"\.0$", "", regex=True)
    )

    receivables = detail[detail["CUENTA"].str.startswith("13")].copy()
    receivables = (
        receivables.groupby(["Identificación", "Tercero"], as_index=False)["SALDO_FINAL"]
        .sum()
        .rename(columns={"SALDO_FINAL": "Saldo"})
    )
    receivables["Saldo"] = receivables["Saldo"].clip(lower=0)
    receivables = receivables.sort_values("Saldo", ascending=False)
    receivables = receivables[receivables["Saldo"].abs() > 0.5]

    payables = detail[
        detail["CUENTA"].str.startswith(("21", "22", "23", "24", "25", "26", "27", "28"))
    ].copy()
    payables = (
        payables.groupby(["Identificación", "Tercero"], as_index=False)["SALDO_FINAL"]
        .sum()
        .rename(columns={"SALDO_FINAL": "Saldo"})
    )
    payables["Saldo"] = payables["Saldo"].abs()
    payables = payables.sort_values("Saldo", ascending=False)
    payables = payables[payables["Saldo"].abs() > 0.5]

    activity = detail.copy()
    activity["Movimiento"] = (
        activity["MOVIMIENTO_DEBITO"].abs() + activity["MOVIMIENTO_CREDITO"].abs()
    )
    activity = (
        activity.groupby(["Identificación", "Tercero"], as_index=False)
        .agg(
            Débitos=("MOVIMIENTO_DEBITO", "sum"),
            Créditos=("MOVIMIENTO_CREDITO", "sum"),
            Movimiento=("Movimiento", "sum"),
        )
        .sort_values("Movimiento", ascending=False)
    )
    activity = activity[activity["Movimiento"].abs() > 0.5]
    return {
        "receivables": receivables.reset_index(drop=True),
        "payables": payables.reset_index(drop=True),
        "activity": activity.reset_index(drop=True),
    }


def _expense_composition(detail: pd.DataFrame) -> pd.DataFrame:
    data = detail.copy()
    data["Grupo"] = data["CODIGO_CUENTA"].astype(str).str[:2]
    labels = {
        "51": "Administración",
        "52": "Ventas",
        "53": "No operacionales",
        "54": "Impuesto de renta",
        "61": "Costo de ventas",
        "62": "Costo de ventas",
        "71": "Costos de producción",
        "72": "Costos de producción",
        "73": "Costos de producción",
    }
    data = data[data["Grupo"].isin(labels)].copy()
    data["Concepto"] = data["Grupo"].map(labels)
    data["Valor"] = data["MOVIMIENTO_DEBITO"] - data["MOVIMIENTO_CREDITO"]
    return (
        data.groupby("Concepto", as_index=False)["Valor"]
        .sum()
        .sort_values("Valor", ascending=False)
        .reset_index(drop=True)
    )


def _statement_preview(
    worksheet,
    values: dict[str, float],
    value_label: str,
) -> pd.DataFrame:
    rows = []
    for row in range(1, worksheet.max_row + 1):
        raw_code = worksheet.cell(row, 3).value
        if raw_code is None:
            continue
        code = str(raw_code).replace(".0", "").strip()
        if not code.isdigit() or len(code) != 4:
            continue
        rows.append(
            {
                "Cuenta": code,
                "Concepto": worksheet.cell(row, 4).value or "Sin descripción",
                value_label: float(values.get(code, 0.0)),
            }
        )
    return pd.DataFrame(rows)


def build_monthly_reports(
    monthly_files: list[BinaryIO],
    previous_file: BinaryIO | None = None,
    initial_balance_file: BinaryIO | None = None,
    start_year: int | None = None,
) -> dict[str, object]:
    if not monthly_files:
        raise ValueError("Debes cargar al menos un balance mensual.")
    if previous_file is None and initial_balance_file is None:
        raise ValueError("Debes cargar el saldo inicial o un informe anterior.")

    workbook, source_name = _load_template(previous_file)
    bce = workbook["BCE"]
    pyg = workbook["P Y G"]
    summary_rows: list[dict[str, object]] = []
    metric_rows: list[dict[str, object]] = []
    monthly_periods: list[tuple[int, int]] = []
    latest_balance_values: dict[str, float] = {}
    latest_pyg_values: dict[str, float] = {}
    latest_raw_df = pd.DataFrame()
    latest_detail = pd.DataFrame()

    if initial_balance_file is not None:
        initial_balance_file.seek(0)
        opening_df, opening_metadata = load_tlg_trial_balance(initial_balance_file)
        opening_detail = prepare_tlg_detail(opening_df)
        opening_year = int(start_year or opening_metadata.get("anio") or 0)
        if opening_year not in PYG_YEAR_TOTALS:
            raise ValueError(
                f"La plantilla actual solo contiene los años {', '.join(map(str, PYG_YEAR_TOTALS))}."
            )
        opening_balance, opening_pyg = _account4_values(opening_detail)
        bce_base_column = _find_bce_month_column(bce, opening_year, 12)
        pyg_base_column = PYG_YEAR_TOTALS[opening_year]
        bce_count = _write_mapped_accounts(bce, bce_base_column, opening_balance)
        pyg_count = _write_mapped_accounts(pyg, pyg_base_column, opening_pyg)
        summary_rows.append(
            {
                "Tipo": "Saldo inicial",
                "Periodo": f"Saldo final {opening_year}",
                "Archivo": getattr(initial_balance_file, "name", "saldo_inicial.xlsx"),
                "Cuentas leídas": int(opening_detail["CODIGO_CUENTA"].nunique()),
                "Filas BCE actualizadas": bce_count,
                "Filas PYG actualizadas": pyg_count,
            }
        )
        metric_rows.append(
            _period_metrics(
                opening_detail,
                opening_year,
                0,
                f"Saldo final {opening_year}",
                "Saldo inicial",
            )
        )
    else:
        opening_year = int(start_year or _infer_start_year(workbook))

    seen_periods: set[tuple[int, int]] = set()
    for uploaded_file in monthly_files:
        uploaded_file.seek(0)
        raw_df, metadata = load_tlg_trial_balance(uploaded_file)
        year, month = _period_from_metadata(metadata)
        if (year, month) in seen_periods:
            raise ValueError(f"Se cargó más de un archivo para {month:02d}/{year}.")
        seen_periods.add((year, month))
        detail = prepare_tlg_detail(raw_df)
        latest_raw_df = raw_df
        latest_detail = detail
        balance_values, pyg_values = _account4_values(detail)
        latest_balance_values = balance_values
        latest_pyg_values = pyg_values
        bce_column = _find_bce_month_column(bce, year, month)
        pyg_column = _find_pyg_month_column(year, month)
        bce_count = _write_mapped_accounts(bce, bce_column, balance_values)
        pyg_count = _write_mapped_accounts(pyg, pyg_column, pyg_values)
        monthly_periods.append((year, month))
        summary_rows.append(
            {
                "Tipo": "Mensual",
                "Periodo": f"{month:02d}/{year}",
                "Archivo": getattr(uploaded_file, "name", f"{month:02d}-{year}.xlsx"),
                "Cuentas leídas": int(detail["CODIGO_CUENTA"].nunique()),
                "Filas BCE actualizadas": bce_count,
                "Filas PYG actualizadas": pyg_count,
            }
        )
        metric_rows.append(
            _period_metrics(detail, year, month, f"{month:02d}/{year}", "Mensual")
        )

    monthly_periods.sort()
    _configure_visible_periods(
        workbook,
        opening_year,
        monthly_periods,
        preserve_existing=previous_file is not None,
    )

    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"

    output = BytesIO()
    workbook.save(output)
    last_year, last_month = monthly_periods[-1]
    metrics = pd.DataFrame(metric_rows)
    third_parties = _third_party_tables(latest_raw_df)
    expense_composition = _expense_composition(latest_detail)
    balance_preview = _statement_preview(
        bce,
        latest_balance_values,
        f"Saldo {last_month:02d}/{last_year}",
    )
    pyg_preview = _statement_preview(
        pyg,
        latest_pyg_values,
        f"Movimiento {last_month:02d}/{last_year}",
    )
    return {
        "output": output.getvalue(),
        "periods": pd.DataFrame(summary_rows),
        "metrics": metrics,
        "balance_preview": balance_preview,
        "pyg_preview": pyg_preview,
        "third_party_receivables": third_parties["receivables"],
        "third_party_payables": third_parties["payables"],
        "third_party_activity": third_parties["activity"],
        "expense_composition": expense_composition,
        "source_name": source_name,
        "start_year": opening_year,
        "last_period": f"{last_month:02d}/{last_year}",
    }


def export_monthly_reports(report: dict[str, object]) -> bytes:
    return bytes(report["output"])
